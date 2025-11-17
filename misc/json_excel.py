import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from dataset_modifier import get_json, file_names


# --------------------------
# Load questions dict
# --------------------------
questions = get_json('data/metadata/questions.json')
# questions is now {"q001": {...}, "q002": {...}}


# --------------------------
# Collect attempts grouped by qid
# --------------------------
attempts_by_qid = {}   # "q001": [attempt1, attempt2]

for file in file_names:
    data = get_json('data/updated/augmented/' + 'train' + '.json')

    for rec in data:
        # extract "q001" from "smp1234q001"
        qid = rec["id"][7:].lower()

        rec["source_file"] = file
        attempts_by_qid.setdefault(qid, []).append(rec)



# --------------------------
# Create workbook
# --------------------------
wb = openpyxl.Workbook()
wb.remove(wb.active)



# --------------------------
# Build sheets
# --------------------------
for qid, attempts in attempts_by_qid.items():

    q = questions[qid]        # master record for this question
    sheet_name = qid[:31]
    ws = wb.create_sheet(sheet_name)


    # ----------------------
    # HEADER
    # ----------------------
    ws["A1"] = "Question ID:"
    ws["B1"] = q["id"]

    ws["A2"] = "Question:"
    ws["B2"] = q["question"]

    ws["A3"] = "Reference Answer:"
    ws["B3"] = q["reference_answer"]

    ws["A4"] = "Max Marks:"
    ws["B4"] = q["max_score"]

    ws["A5"] = "Total Attempts:"
    ws["B5"] = len(attempts)

    ws["A6"] = "Answer Breakdown:"
    ws["B6"] = (
        f"Correct: {q['count']['correct']}, "
        f"Partially Correct: {q['count']['partially correct']}, "
        f"Incorrect: {q['count']['incorrect']}"
    )


    # bold left header column
    for r in range(1, 7):
        ws[f"A{r}"].font = Font(bold=True)

    # wrap long text
    ws["B2"].alignment = Alignment(wrap_text=True)
    ws["B3"].alignment = Alignment(wrap_text=True)
    ws["B6"].alignment = Alignment(wrap_text=True)



    # ----------------------
    # TABLE HEADER
    # ----------------------
    start_row = 8
    headers = ["Provided Answer", "Normalized Score", "Verification Feedback", "Source File"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(start_row, col, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 30



    # ----------------------
    # GROUP ATTEMPTS
    # ----------------------
    grouped = {
        "correct": [],
        "partially correct": [],
        "incorrect": []
    }

    for a in attempts:
        grouped[a["verification_feedback"].lower()].append(a)



    # ----------------------
    # WRITE ATTEMPTS
    # ----------------------
    row = start_row + 1

    for group in ["correct", "partially correct", "incorrect"]:
        for a in grouped[group]:
            ws.cell(row, 1, a["provided_answer"]).alignment = Alignment(wrap_text=True)
            ws.cell(row, 2, a["normalized_score"])
            ws.cell(row, 3, a["verification_feedback"])
            ws.cell(row, 4, a["source_file"])
            row += 1

        row += 1   # blank line after each group



# --------------------------
# SAVE EXCEL
# --------------------------

sorted_names = sorted(wb.sheetnames)
wb._sheets = [wb[name] for name in sorted_names]
wb.save("data/metadata/augmented_samples.xlsx")